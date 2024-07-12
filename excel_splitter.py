import sys
from PySide6.QtWidgets import (QApplication, QWidget, QVBoxLayout, QPushButton, QLineEdit, QLabel, QFileDialog, QMessageBox, QComboBox)
from openpyxl import load_workbook, Workbook
import os
from PySide6.QtGui import QPixmap, QIcon
from PySide6.QtCore import QByteArray, Qt, QTimer


class ExcelSplitterApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Excel Splitter')
        self.setGeometry(100, 100, 400, 250)
        base64_image = b"iVBORw0KGgoAAAANSUhEUgAAADIAAAAwCAYAAABT9ym6AAAABGdBTUEAALGOfPtRkwAAACBjSFJNAACHDwAAjA8AAP1SAACBQAAAfXkAAOmLAAA85QAAGcxzPIV3AAAKL2lDQ1BJQ0MgUHJvZmlsZQAASMedlndUVNcWh8+9d3qhzTDSGXqTLjCA9C4gHQRRGGYGGMoAwwxNbIioQEQREQFFkKCAAaOhSKyIYiEoqGAPSBBQYjCKqKhkRtZKfHl57+Xl98e939pn73P32XuftS4AJE8fLi8FlgIgmSfgB3o401eFR9Cx/QAGeIABpgAwWempvkHuwUAkLzcXerrICfyL3gwBSPy+ZejpT6eD/0/SrFS+AADIX8TmbE46S8T5Ik7KFKSK7TMipsYkihlGiZkvSlDEcmKOW+Sln30W2VHM7GQeW8TinFPZyWwx94h4e4aQI2LER8QFGVxOpohvi1gzSZjMFfFbcWwyh5kOAIoktgs4rHgRm4iYxA8OdBHxcgBwpLgvOOYLFnCyBOJDuaSkZvO5cfECui5Lj25qbc2ge3IykzgCgaE/k5XI5LPpLinJqUxeNgCLZ/4sGXFt6aIiW5paW1oamhmZflGo/7r4NyXu7SK9CvjcM4jW94ftr/xS6gBgzIpqs+sPW8x+ADq2AiB3/w+b5iEAJEV9a7/xxXlo4nmJFwhSbYyNMzMzjbgclpG4oL/rfzr8DX3xPSPxdr+Xh+7KiWUKkwR0cd1YKUkpQj49PZXJ4tAN/zzE/zjwr/NYGsiJ5fA5PFFEqGjKuLw4Ubt5bK6Am8Kjc3n/qYn/MOxPWpxrkSj1nwA1yghI3aAC5Oc+gKIQARJ5UNz13/vmgw8F4psXpjqxOPefBf37rnCJ+JHOjfsc5xIYTGcJ+RmLa+JrCdCAACQBFcgDFaABdIEhMANWwBY4AjewAviBYBAO1gIWiAfJgA8yQS7YDApAEdgF9oJKUAPqQSNoASdABzgNLoDL4Dq4Ce6AB2AEjIPnYAa8AfMQBGEhMkSB5CFVSAsygMwgBmQPuUE+UCAUDkVDcRAPEkK50BaoCCqFKqFaqBH6FjoFXYCuQgPQPWgUmoJ+hd7DCEyCqbAyrA0bwwzYCfaGg+E1cBycBufA+fBOuAKug4/B7fAF+Dp8Bx6Bn8OzCECICA1RQwwRBuKC+CERSCzCRzYghUg5Uoe0IF1IL3ILGUGmkXcoDIqCoqMMUbYoT1QIioVKQ21AFaMqUUdR7age1C3UKGoG9QlNRiuhDdA2aC/0KnQcOhNdgC5HN6Db0JfQd9Dj6DcYDIaG0cFYYTwx4ZgEzDpMMeYAphVzHjOAGcPMYrFYeawB1g7rh2ViBdgC7H7sMew57CB2HPsWR8Sp4sxw7rgIHA+XhyvHNeHO4gZxE7h5vBReC2+D98Oz8dn4Enw9vgt/Az+OnydIE3QIdoRgQgJhM6GC0EK4RHhIeEUkEtWJ1sQAIpe4iVhBPE68QhwlviPJkPRJLqRIkpC0k3SEdJ50j/SKTCZrkx3JEWQBeSe5kXyR/Jj8VoIiYSThJcGW2ChRJdEuMSjxQhIvqSXpJLlWMkeyXPKk5A3JaSm8lLaUixRTaoNUldQpqWGpWWmKtKm0n3SydLF0k/RV6UkZrIy2jJsMWyZf5rDMRZkxCkLRoLhQWJQtlHrKJco4FUPVoXpRE6hF1G+o/dQZWRnZZbKhslmyVbJnZEdoCE2b5kVLopXQTtCGaO+XKC9xWsJZsmNJy5LBJXNyinKOchy5QrlWuTty7+Xp8m7yifK75TvkHymgFPQVAhQyFQ4qXFKYVqQq2iqyFAsVTyjeV4KV9JUCldYpHVbqU5pVVlH2UE5V3q98UXlahabiqJKgUqZyVmVKlaJqr8pVLVM9p/qMLkt3oifRK+g99Bk1JTVPNaFarVq/2ry6jnqIep56q/ojDYIGQyNWo0yjW2NGU1XTVzNXs1nzvhZei6EVr7VPq1drTltHO0x7m3aH9qSOnI6XTo5Os85DXbKug26abp3ubT2MHkMvUe+A3k19WN9CP16/Sv+GAWxgacA1OGAwsBS91Hopb2nd0mFDkqGTYYZhs+GoEc3IxyjPqMPohbGmcYTxbuNe408mFiZJJvUmD0xlTFeY5pl2mf5qpm/GMqsyu21ONnc332jeaf5ymcEyzrKDy+5aUCx8LbZZdFt8tLSy5Fu2WE5ZaVpFW1VbDTOoDH9GMeOKNdra2Xqj9WnrdzaWNgKbEza/2BraJto22U4u11nOWV6/fMxO3Y5pV2s3Yk+3j7Y/ZD/ioObAdKhzeOKo4ch2bHCccNJzSnA65vTC2cSZ79zmPOdi47Le5bwr4urhWuja7ybjFuJW6fbYXd09zr3ZfcbDwmOdx3lPtKe3527PYS9lL5ZXo9fMCqsV61f0eJO8g7wrvZ/46Pvwfbp8Yd8Vvnt8H67UWslb2eEH/Lz89vg98tfxT/P/PgAT4B9QFfA00DQwN7A3iBIUFdQU9CbYObgk+EGIbogwpDtUMjQytDF0Lsw1rDRsZJXxqvWrrocrhHPDOyOwEaERDRGzq91W7109HmkRWRA5tEZnTdaaq2sV1iatPRMlGcWMOhmNjg6Lbor+wPRj1jFnY7xiqmNmWC6sfaznbEd2GXuKY8cp5UzE2sWWxk7G2cXtiZuKd4gvj5/munAruS8TPBNqEuYS/RKPJC4khSW1JuOSo5NP8WR4ibyeFJWUrJSBVIPUgtSRNJu0vWkzfG9+QzqUvia9U0AV/Uz1CXWFW4WjGfYZVRlvM0MzT2ZJZ/Gy+rL1s3dkT+S453y9DrWOta47Vy13c+7oeqf1tRugDTEbujdqbMzfOL7JY9PRzYTNiZt/yDPJK817vSVsS1e+cv6m/LGtHlubCyQK+AXD22y31WxHbedu799hvmP/jk+F7MJrRSZF5UUfilnF174y/ariq4WdsTv7SyxLDu7C7OLtGtrtsPtoqXRpTunYHt897WX0ssKy13uj9l4tX1Zes4+wT7hvpMKnonO/5v5d+z9UxlfeqXKuaq1Wqt5RPXeAfWDwoOPBlhrlmqKa94e4h+7WetS212nXlR/GHM44/LQ+tL73a8bXjQ0KDUUNH4/wjowcDTza02jV2Nik1FTSDDcLm6eORR67+Y3rN50thi21rbTWouPguPD4s2+jvx064X2i+yTjZMt3Wt9Vt1HaCtuh9uz2mY74jpHO8M6BUytOdXfZdrV9b/T9kdNqp6vOyJ4pOUs4m3924VzOudnzqeenL8RdGOuO6n5wcdXF2z0BPf2XvC9duex++WKvU++5K3ZXTl+1uXrqGuNax3XL6+19Fn1tP1j80NZv2d9+w+pG503rm10DywfODjoMXrjleuvyba/b1++svDMwFDJ0dzhyeOQu++7kvaR7L+9n3J9/sOkh+mHhI6lH5Y+VHtf9qPdj64jlyJlR19G+J0FPHoyxxp7/lP7Th/H8p+Sn5ROqE42TZpOnp9ynbj5b/Wz8eerz+emCn6V/rn6h++K7Xxx/6ZtZNTP+kv9y4dfiV/Kvjrxe9rp71n/28ZvkN/NzhW/l3x59x3jX+z7s/cR85gfsh4qPeh+7Pnl/eriQvLDwG/eE8/s3BCkeAAAACXBIWXMAAFxEAABcRAFraoDIAAAAIXRFWHRDcmVhdGlvbiBUaW1lADIwMjM6MTE6MDMgMTI6NDI6MDGUwlLtAAAMNklEQVRoQ+1Ze3AU9R3/7N7tvR8JSS4PQkISSEIQCOERBZEmKKNTwApqtba1WgRL7ehMpzO2U63Y6dhai1UHER1HW4qltj5GVKpQlQFkBJQhBEQeAYkJhISQxz327nZv+/lt7oAEEppMtP/wmfnd3u1v97ffz/f9u8VlXMZlDAgpefx6sHqJgsxQJuyWclhkG7JsH2Dq8/Hk7LBi+Ii8eosHhm0kXJYyJBLl0I2JsMoliCWKIUsBeOxAUP05Fq5dkbxjWDF4Iq/eYoHFngGrRQhYhpheAYtUwZVKENcKoFjdUGTA4LV6AohzaBw+EonE12Dh337Ys9DwYmAiby1xQVNzYUcpNAqtaVdQy6VIGMXQjJFwWgGZQguBU0InOASJvvCSSFR/EQvXLE6eGVZcSOS9xZXQ4oug6hOgSEV0jSIK74XNck7LQsNiJC4mcT/wOQA1shoL192bPDOs6E3kldvy4FJ20Z9zoTImUwLrFNgYhNAXg8uKDMO5rm3O6j80n272qMnTg4HVCkmRbOG8vKxPk6fOojeRN++qQDy2m2dtJoHhRCyIO0vmRl+qWmbt0lSLMQTFSBIFs1qMhJ54fufOT+6rqanRklOgg58PSSeJWPLHMMNgDrDYJVkeEgkB8z7ZImmGsbSsrGxs8rSJPkS+Xui6DoOWFsehjng8DiORiDl8PltyWRNDJyK0YzCOzPG/adgqyZAsMhwW5eyQ6S7nwyZb4bTYes33XZ0ulqB1evn+IInwoUJoKQq728AYXw5KOBSXWDM6MCGS6NZiaAq3ozlyBk0czeoZkrMkhTVMwbu0MI5H2vBVhNepHUycVpapC5NrX/QJ9rtZK6K7+M1zQbALBcgxuN12TDQKkbunANGdtC5XkKeGcWLScdSjEWqI8WeI830eTo4O2QIftS3oiiFpKr476hr8fsKdIn7wYete/PSzZ3EmoZEcradF8YPCOfjd+DsQ1mMmWUVRhGupumRMz/B695prE5e2iCAgqfB6Jcx0lOGm/bUIPFcOaasfM6qKUXNVKfL2j0LBixOw8FAtqt0lcHl6rNbLQrIEldo+1d2E1mAz2jhauxvxUVs9s3uCLqWgvvM4jpw+gPbgSc6fMOc3i3kS6OuCfXEJi8Tg89hQqRWjcH8RojvsOHa4AyXj0/H0yu8gGIxB13SMLhqB5Q9vxIbXD5lzSnUYDWVHsEf+EqGg6BFZ1fUoJqeXYEHOZMQZtIKixdAxN2cKqtLYKCR0hHjNK41bcDrWRYPKsFIR1+dOxST/aNbnnl6zP4v0TyQRw0R5FCbVX4HwpwpaT4Sg2eIIRmJYeNMV+M1Dc7FzVyM2bTrIVSTsqz+JhmOnWUAl2BNsdPNdsE9WUVd+AHWJ4wyhMJaNXYCV05eY4WSCT1fpPkG6mICIlzSbx/yeQoTzIV3lpT2iDs61qAlJ0VHeVYjQBhsav+yA5ojDapfh89qwdetRPPXUFjgcCh544BrcvGgiXG6FnU0CikNG1BrFV4e6oH3kRJEaoITCurLp51F2DGeiQbRHu9HBo5k7kkIKImEK3sHi2RkLoSseNotgan4g9G+RWAyVIwoxJZoPdZuM9s+c6O7U0M3FF908AfPnjcd773+BmtljMH36KKxctRUrn9mOgMcPf64E/4wQ5Ol21Out2N1ajwRdJYPaLnSMMGNCuJZMYrcXzMbPSr7N3xI+Z0z8sn4N2kwCzGb0ijtH1+LHhdchcolgvziRRNxTYPOj2P8I2hgnY907kX7qOGI7rGjd7oQeseDun1Rh+pUFLFIJ7Pi4EX95fjccWTrSZnXDqLKizVOGw53VyLM2kshyBBOME2ob9P+ejMYRD2JM5njsqn0Sfrsbj33+L/xqxxMMKV/PfKwbFdmTsaP2CTbWLIhUwOCI6FHPOFcOJNuj2N/pZg8po9DXjCL3bgQ6jkLdnkDzZhssIYfoGGCkR5BZE4Flmg0tjnI0dFahMZjOh8uYlvYZDrQ/RktKKPbkojqN2xYGuWicJRbTW0bOxPzcafxtsHa04Y8H38BJErAwHUtMw99jel6QW02LRE0rDppImTMbdsdy1HWzJoiMofMo25Hna0NlRh3Suvah/YBKIgZ8FU4SmIj6tkq0hvyMsQhXZj2xelDtr8P+0yQSUbF07Hw8N/VeaDGdDxPOJHNbQ3fVeD1hZwp2Wuwk1VPDRHyI791MBilBBxfsArxTaIBJHPNG2vDyVQqemZLAKFsa3j02B3vlH6F5XAdOjfNgd/QebD/+LVyX4cX6qw1su9aFe0pEyuWWJbVYLwhvT32eg/gt6oUIejEsHD1X9L7uYuAW7+IQj1a5DxnllrF+hpttRQKftGnYUuPA/XsiWHUoHXMD09Ch5+CLdhf+NDmG+0ocePxgFHlOalswEMowZeCH1Yn3W/fhtu0rTH8392SJOG7Nn4kb86pN/z8V7cTy/evQIlxLEKELLsyfgdqsiWawD4QBK7uV2jkTNbCVBDxWkVUSeLAugtMR9kWKxs2iTqHikK06Zmcp3ERKWJCnIMsms1Jz73D+6mxNjnY34x9HNuD1Yxvx5pccR97BQ/vWmjHjVVx4teljrNjzAtY2vIu/HnkbLx34p5nFBGcLs95AGHiWSLdJWHFQxarDUXx/tA1Li+1ojyUg3FwkH/Iz3eeuXWE8uDeCujMa5udZ8UyVCxVeCzXJSXEhAzfTmY4puVMwOVCJSo7JrNpLS26AjY2hCOZZGRWozr8aZemlGJdRjvGBiWawi6ZRJIOBIDzoHM4L9gp3HiLyI3DKNmya7cR/WuLoonWXjbXjreY4btxmYE7W68z5AWaZWVg7TbgimOV0/KLcgTrWnFkfGhjt2Y9jHb9FV1jFXSXXY3XVvQjHVJO8KHSi4w0xmEV8OGg14WIq3UjMiXhxWx1mdU/RGFyw04wRNnjFLg2fd7mwYGsEPvrJnBwrdrRr+PMhZiveaeGlCh/WrhrYdEpDKS0wP1/BqoYobtgSpygKsm2s8OKflaSgit0KPwtjOkcaa4dAKuhF7PgVN7JpuYAzDZkOP2PMMJvGS+HiFmFll2NRTPCNRrZ7Hg6Hq9AQdjI9qtAQN/9IcSgeXON/DR2JbOzonMkMFYKV5DT2WVbZgfGeNgSUD3CgYxMa1SCl1VHJ5nBeTiXvZ2WngDK1L5rGqWwmBYkYU/H65p1oY7CLppG7c8zJrsQYTw6ies/2fHB1JNX96hF4LDrGeksxwnktWuJX4lDYx0XDsDLf16a9wawVIJEZFJQJQXGgzN0Cv/QhC+IHOBI6SRdy0u49+xaI5pA91NmnshWpYEbaXvM4fAz2pw+/jftFZWeGM+MqHsKMvCvx/tWPIsakImrK4OuIgMWFoME60XEE21pWQoo9jGt861HljyBgUxiEBrVvRZZiY8XuQLXn7+gK/RoftbyCQ8EOJGQWRxY5M/3SCk66TY5/FLK9+QhwZPsKcW3WJLNmaEwGo12ZyOV5pyMdbvZkHmcWvaKAQpLUkIL97H6kD9hOK1IMRZ4C5LmnoqF7DzWZjYAjH8e6N+NouIXPc1ABLIZ9/Zp91e2FtXiCu8FI/FywZ9v9TL+aqW2xVz/BLXAn41NUfRHsha4s8/ZUtR+aa/UHdqUw2ABKFNoUSRQrfhcu1F9g0u8XM9W+ULUMoVhPSyIg4uL81OqgBUXNSJ2J0qU0kkgJOjTX6g9CYIu3x23Yf0FOfr9EdjHbd7qYSuFSo299EOfETjGcHOKe3to2ISKo1+mhERkizE0S87Z5HMLgB9M+k74k2aMxVuXz0IeIwS7NoO2SP4cTXNMUhh7o83qGNPw+LxOATQj97yZVbUiubKK31V67I5+Ud8Nrz2RE9vyBLUZ/rwoGA26BxyiZ7xyc++TLdB+H+NdwsKAtZNafUFNT08bS0lLu0M6hNxGBDYtnIR6/lc5bRumLSKSA3aCocj1xnXqtII6Dfa0QiazConXLkmeGFRcSOR8bl/jRHixgpStl6mBGS4gXPWN4LOKdmWC7Ya4gXjtcynrf+IueS+Gp6+3Iz8iCzTqGFiplbzGexaWMsVXM8lvAhspuWk+QSZET4//66m0wWHOHDz7LKG62y6BrtB4msGSXkEQRH5MFP12rS32ERJYn7xhWDB+Ri0G8nk5rD8DlHsvmzA2HYzNqnmUHeRmXcRnfLID/Ahzy3wJ6zxYDAAAAAElFTkSuQmCC"
        pixmap = QPixmap()
        pixmap.loadFromData(QByteArray.fromBase64(base64_image))
        self.setWindowIcon(QIcon(pixmap))

        layout = QVBoxLayout()

        self.input_file_label = QLabel('Input Excel File:', self)
        layout.addWidget(self.input_file_label)

        self.input_file_path = QLineEdit(self)
        layout.addWidget(self.input_file_path)

        self.browse_button = QPushButton('Browse', self)
        self.browse_button.clicked.connect(self.browseFile)
        layout.addWidget(self.browse_button)

        self.sheet1_label = QLabel('Select Sheet 1:', self)
        layout.addWidget(self.sheet1_label)

        self.sheet1_combo = QComboBox(self)
        self.sheet1_combo.currentIndexChanged.connect(self.loadColumnNames)
        layout.addWidget(self.sheet1_combo)

        self.sheet2_label = QLabel('Select Sheet 2:', self)
        layout.addWidget(self.sheet2_label)

        self.sheet2_combo = QComboBox(self)
        layout.addWidget(self.sheet2_combo)

        self.column_name_label = QLabel('Select Unique id column to split rows:', self)
        layout.addWidget(self.column_name_label)

        self.column_name_combo = QComboBox(self)
        layout.addWidget(self.column_name_combo)

        self.num_rows_label = QLabel('Number of Rows to Split:', self)
        layout.addWidget(self.num_rows_label)

        self.num_rows_input = QLineEdit(self)
        layout.addWidget(self.num_rows_input)

        self.split_button = QPushButton('Split', self)
        self.split_button.clicked.connect(self.splitSheets)
        layout.addWidget(self.split_button)

        self.setLayout(layout)
        self.setStyleSheet("""
            QMainWindow {
                background-color:#f5f5f5; /* White background */
            }
            QLineEdit {
                background-color: white;
                border: 1px solid #ccc; /* Light Gray border */
                padding: 4px;
                width: 150px; /* Adjust as needed */
            }
            QPushButton {
                background-color: #0d6efd; /* Dark Blue button */
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 3px;
                width: 150px;
            }
            QPushButton:hover {
                background-color: #0951ba; /* Slightly darker grey on hover */
            }
        """)

    def browseFile(self):
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getOpenFileName(self, "QFileDialog.getOpenFileName()", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if fileName:
            self.input_file_path.setText(fileName)
            self.loadSheetNames(fileName)

    def loadSheetNames(self, fileName):
        wb = load_workbook(fileName, read_only=True)
        self.sheet1_combo.clear()
        self.sheet2_combo.clear()
        for sheet in wb.sheetnames:
            self.sheet1_combo.addItem(sheet)
            self.sheet2_combo.addItem(sheet)

    def loadColumnNames(self):
        fileName = self.input_file_path.text()
        sheet1_name = self.sheet1_combo.currentText()
        if not fileName or not sheet1_name:
            return

        wb = load_workbook(fileName, read_only=True)
        sheet1 = wb[sheet1_name]
        self.column_name_combo.clear()

        # Use iter_rows to extract column headers
        first_row = next(sheet1.iter_rows(min_row=1, max_row=1, values_only=True))
        for cell_value in first_row:
            self.column_name_combo.addItem(cell_value)

    def splitSheets(self):
        input_file = self.input_file_path.text()
        sheet1_name = self.sheet1_combo.currentText()
        sheet2_name = self.sheet2_combo.currentText()
        column_name = self.column_name_combo.currentText()
        num_rows = int(self.num_rows_input.text())
        output_file = os.path.splitext(input_file)[0] + '_split.xlsx'

        if not input_file or not sheet1_name or not sheet2_name or not column_name or not num_rows:
            QMessageBox.warning(self, 'Error', 'All fields must be filled out.')
            return

        try:
            self.split_and_create(input_file, output_file, sheet1_name, sheet2_name, column_name, num_rows)
            QMessageBox.information(self, 'Success', f'Sheets split and saved to {output_file}')
        except Exception as e:
            QMessageBox.critical(self, 'Error', str(e))

    def split_and_create(self, input_file, output_file, sheet1_name, sheet2_name, column_name, num_rows):
        wb = load_workbook(input_file)

        # Load the sheets
        sheet1 = wb[sheet1_name]
        sheet2 = wb[sheet2_name]

        # Find the column index based on the column name in sheet1
        col_index_sheet1 = None
        for col in sheet1.iter_cols(1, sheet1.max_column):
            if col[0].value == column_name:
                col_index_sheet1 = col[0].column
                break
        if col_index_sheet1 is None:
            raise ValueError(f"Column {column_name} not found in {sheet1_name}")

        # Find the column index based on the column name in sheet2
        col_index_sheet2 = None
        for col in sheet2.iter_cols(1, sheet2.max_column):
            if col[0].value == column_name:
                col_index_sheet2 = col[0].column
                break
        if col_index_sheet2 is None:
            raise ValueError(f"Column {column_name} not found in {sheet2_name}")

        # Split Sheet1 based on the number of rows
        rows_to_split = [row for row in sheet1.iter_rows(min_row=1, max_row=num_rows)]
        if len(rows_to_split) < num_rows:
            raise ValueError(f"The sheet {sheet1_name} does not contain {num_rows} rows.")

        # Create a new workbook for the output
        new_wb = Workbook()
        new_sheet1 = new_wb.active
        new_sheet1.title = 'Sheet1'

        # Add rows to new sheet 1
        for row in rows_to_split:
            new_sheet1.append([cell.value for cell in row])

        # Gather corresponding data from Sheet2 for the selected rows based on the common column
        values_in_column = {row[col_index_sheet1 - 1].value for row in rows_to_split if len(row) >= col_index_sheet1 and row[col_index_sheet1 - 1].value is not None}

        # Create a new sheet for the second output (Course sheet)
        new_sheet2 = new_wb.create_sheet(title='Sheet2')

        # Filter and add rows from sheet2 to new_sheet2 based on values_in_column
        for row in sheet2.iter_rows(min_row=1, max_row=sheet2.max_row):
            if row[col_index_sheet2 - 1].value in values_in_column:
                new_sheet2.append([cell.value for cell in row])

        # Save the workbook with both sheets
        new_wb.save(output_file)

def main():
    app = QApplication(sys.argv)
    ex = ExcelSplitterApp()
    ex.show()
    sys.exit(app.exec())

if __name__ == '__main__':
    main()
