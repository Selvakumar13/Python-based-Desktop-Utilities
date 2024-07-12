import sys
import openpyxl 
from PySide6.QtGui import QPixmap, QIcon, QGuiApplication
from PySide6.QtCore import QByteArray, Qt, QTimer
from PySide6.QtWidgets import QMainWindow,QMessageBox, QApplication, QVBoxLayout, QCheckBox, QLabel
from PySide6.QtWidgets import QTextEdit, QPushButton, QWidget, QTableWidget, QFileDialog, QHBoxLayout
import subprocess

base64_image = b"iVBORw0KGgoAAAANSUhEUgAAAWsAAAEfCAYAAACH9eevAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAABprSURBVHhe7d1tiCVXncfxmqeenu7pycwk85CooEYWhMRlBVkT8IWLcREJyb5QwYQNKKPEN+oaFFmXsKvgKj6yYDYbFFySgFnEiIiYuO4Ll01EUEh8tWxU0DzM80NPz1PmYe/vTtWkUlN1b51/nao6p+r7gaZvdyZ9695b9Tun/nXOqQQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAgMFal37v0qX0OwDEqvPs7OoJCWgAQ9VJjrb9JIQ0gLFoNU/b/OMENYAxaiVX2wprghrAmHnPVt9/kJAGgMu85uv69DsAwC+vnVefyU+vGgCu5iVnfYU1QQ0A1RpnrY+wJqgBYL5GeUvNGgAi0LRnTa8aAOozZy49awCIAGENABFoUgahBAIA7ky5S88aACJAWANABAhrAIgAYQ0AESCsASAChDUARICwBoAIENYAEAHCGgAiQFgDQAQIawCIAGENABEgrAEgAoQ1AESAsAaACBDWABABwhoAIkBYA0AECGsAiABhDQARGOQNcx+97vrkLZs2pz/V88jaieSLJ46kP2Go3rtlObltcXm6f+zesCH9bZI88/LZ5HfnX04eW1udPu6atutL23elP9X3mWMHkx+fXkt/QiS4YS5Q5YYNG6eNuALxXYtLrwpqUXjfuWXr9N98Yft16W+BcBDWGDwF9YM799Q+21Jof2PH7vQnIAyENQbvnuVtyRs2bkp/qke9b5UmgFAQ1hg81agtrP8f0AbCGlH5wNLKtKb8892vS38zX7E+XZf1/wPaQFgjaO/YvCX5+MqO5KGde5LfXv/65B+uuXZaUyZIMTaENYKii4HqPesCn3rPD0xCet/Wa5JbJqFtdeDChfSRGw3lA0JBWCMIn922M/nRrtckT+x+7bT3XDa8zuqRUyfSR26eOns6fQT0j7BGEO4yjNio69snjztPdHn89EkmmyAohDVG4b6jB6cBXIdms37u2KH0JyAMhDVG4YUL56cBrOnZPztzKvl9oR6tnrfC/IOHXmTZAQSJsMaoqLTxiaMHktsPPp/c9OIfrnwppBXmfawLAtRBWANABFh1LxXTqnt6be9fXkn+fPI9uyin4Wn3Hz+U/CLSEQwaQ+1KPeIQafjhe7YsJ29fWEz2Th7nL5yq/PLc5EsjTb53ajX9bX+r7mlfetvmxeTmyXeNvikeN2uXLk63V8MYte0/mTyXSkpoxJS7hHWqKqz1d/T3LLRz63Q742O7NMRNIyfK5A/cWf9ullkBaNl+lRVUYshYQ8mFr+d0bQwU0p+evO8adliHGtgHTx6bhnbXYa3n0/7h+nmKav7fn2zzvI6BZppqApOruq/J0sCrkdx3ZH/6U29YIrUN6sFa/evkQPRJE0UsAYz2fXjrNckPdt1QO6hFPVmNKe9yhT/NCFWjq4bBEtSi16jJSgpjNVBVfmrs8b9p40L6qJoaG4unz51JH8WHsJ5BO+JtDgdfnnofTU5Pi3RguAQBuqOw/eTKjmR5ne1w0udq6YG60rR9haw1pIu0zVp6Vg1AGfW8LbNHb9o0P6zrBHoZlXFiRVjPoLqj5QBUne/LHuvf6kV0cTDDna9GtMl0+jrUoGjavm+qx39lx67KwP6fc+7XUG6sEcR1Ar1IHaiY6+2E9Qx3LdlKDo+fOul1p6D0ESaVPmJoRBXUbZ6VqUOjWn1ZScRSCim70FlUJ9CLYl8+gLCuoN6sZW0Knfb5HFXyxknPxddpK/zRZ/KRFnqqvqlB6aJ8ph72x1a2pz+9wloKefOMnrMaBddjU2e7+dE3MSKsK1h7TF9d9Tv8r+3TY9h8aBKC1hp1V7puUHTMlHUsLKWQmxeqOyh/MeO/VXnqbLwXFjOEdQntcJaQ1GmWz4uKCJP2jxgu9loaFE25f/eBP02HLd57ZL/zjE49Z5GlFKIzyiplDcI8Q1hBkbAuYRmup9Osf1n1O1QPYWoynLMrlgZFQa0p99n1FpUwNF69uI7KLLdsXryqdm0phcwK5FlBXmYIJRAhrAusw/V0UZF1JebTmYd6bcUvi7K/o6/8hJg23LpgL01pkpO2L9tO/aww8c11HLK24VsVnY2HJ9tYl3ryZSNDLKWQqtfgenHxyTOn0kdxI6wL3re04nzq6PuiYhU9z9dXj1452PWlU1X9bgineTFQEFkuPIs+J+0nWaOu7/r5n44fnv7s062OZTwFWtUIJtdeaVm92VIKKRtLbbm4aJ2cExrCuuAOw4VFTRlum8aI/tWBP161kL5OMfU7TaGlXt6+ty4spo/c6PPT51RGn5v+uy8qIbjeyOHZc7PPCl3KGGVlCksppGwsddV47ip6Tj33EBDWObr3n2urreBsux6m59CynuifawhmvlMR1BmtteHLrGFvVU7OKcW8dLH+vIG968unoLuWQsrKHa71akv5JVSEdc4dS+696n8+3n75Q3VNhME69j5/NlTGZ+/PNdBEa4VoYaSqL5cRGFXvkWs5Qn+n+LyzhvSVGUoJRAjr1J4NG512SNHV83kHYVM60ClvhGPFMLbapVc6ZJZSSPEs4UaHhmhIJRAhrFOWcbNNRgXU9dz5c+kjhMBSBplXD/bNtffZJdeyRP61qDPlcvH/hzXvuRkLwjplucCj0zRN522TFn0HhsK1LJEv6bjW4v/L40XbEBDWqf0XzpuGv2k6b9kCNkBmeT2HWca1FJIvTbrU4jWRp+0SZdfYi3JUg3al07J7WBVvNFxrrlInZFyvl8xi2cYuuZZCsskxLuUdn0MhQ0FY5+hCnsvU2sydS1vpXfegj/fccrGwzkUx3QfRF50lhsy1FJJNjnG5uDi0EogQ1gVj711X9fAUjD57fz5YVl9rau2i+9Rw7R8awz+Lde30MgcuuvesP3/88JVZsT6+ZnEtheiirsvFRZU/hlYCEcK6QLPMLKeR6l2HFGaWA1aq1mMIsTG6bdF2H74mrBd87568f1VnArrdlmX8dpVfGZYDnbV9bXAphei9cbm4+J8D7FULYV3iyTPu45rV6pctD9mXl4ynwrorje6Mnh24aoCsd0pvm4Zb5re1C9Yem3qHul9hvoetqdNt3G5L2+ja4ci2z7XDofdejY3rTX9dSiHaJpeLizHfZ3EWwrrEd40roSk8Quld//G8vW6pYH5i92unM9d0J+wugtq68lx+W7MvbXNbftNgzLQCUXczz7ZTN69ta11syzRrbZ/eOwWvhqSWrcOh3+nsS42k/q3eezU2rmcGrqWQuhcXNaIr5vsszkJYl9CHbb2zRCi9a/Wu2lh6sy3PRTKe/PK+4R6EXXtszb7WiBoQ3a1djUm+EcwaGE1NVyPZtGPi0qDUfa6nz8V/R5gqhHWFeQvvVNGO7royWFue6XjmXBO/jOi2S5aL0F1TYx36draxbsdQSyBCWFfQzm7tQd1bcuPQPsTUy4hpqJWGeFpr113SzQRCHnNtWStkFo2tHmoJRAjrGaw9E52yud6pow3qZcRSClH4xTSR4YEIbuGm4Lr/+KGg9wFLbb3K0G/AQVjPYJ0kIyGMntDB+miD2mWewrTt3qRKT7E0LuoVPmQslZVpq2Sh7bzv6MFg31dfpRC9viHcZ3EWwnqO2HvX31w92jhk1WDpgG+btvNrJ46mP4VP762Ps4G2b8uWBXaIJRFfpRDrgICYENZzWCfJyKdWdqaP+qUD1XqGoBD56JH9ndUC1TvSbLpYeti6g4/C1rK9+n/0/1bd7ssnheLdhy/foDc0PkohQy+BCGFdg3Vd3C6WUK1DQXv7weedDlQ1UAoS3dux64s2Cux9h/dHU8Oe3gPTcXv1b/X/dBHUGX2OukFvdld1Hz1anQ2pHNTkzKtpKWQMJRBZl363uJR+R0Q040x3cP/LzYvTe+XlJzOo963xzuqlhLLzq5z0zsWl6c1T9062XRM38nSgapu1Zoemgis8+ryzTn57dQ/B/PurbdM2agx009KULxpmqpsA633VtmqxpLI1OLRvrE7eawW8ForS9muCUNcN+UCYcpewBoBumXKXMggARICwBoAIENYAEAHCGgAiQFgDQAQIawCIAGENABEgrAEgAoQ1AESAsAaACBDWABABwhoAIkBYA0AECGsAiABhDQARIKwBIAKENQBEgLAGgAgQ1gAQAcIaACJAWANABAhrAIgAYQ0AESCsASAChDUARICwBoAIENYAEAHCGgAiQFgDQAQIawCIAGENABEgrAEgAuvS7xaX0u/AOPz7S+mD1N/uTR8ATky5S1gDdRSDehZCHLMR1kArXIK6CgGOVxDWQCt8hHUVQnyMCGugFW2GdRkCfOgIa6AVXYd1FUJ8KAhreBZKSGE2Qjw2hDU8IqjjRoCHjLCGR4T1MBHiISCs4RFhPS6EeJcIa3hEWIMAbwthDY8Ia1QhxJsirOERYY26CG9Xptxl1T0AzdCwd4KwBoAIENYAEAHCGkAz1Kw7wQVGlKMOiSqEc1OMBoFHhDUI5bYQ1vCIsB4XgrlLhDU8IqyHiVAOAWENjwjr+BHMoSKs4RmBHQdCOTaENdCKUBotQnkoCGugFX2ENcE8ZIQ10Io2w5pQHiPCGmiFj7AmlPEKwhpojUtgE8yYjbAGWlUMbEIZNoQ1AETAlLusugcAESCsASAChDUARICwBoAIcIERw8KIDYSP0SAYmbpjnwlshIWwxoA1mUVIWCMshDUGokkwlyGsERbCGhHyHcxlCGuEhbBG4LoI5jKENcJCWCMQfYUyytFYhYawRg8I5jgQ2CEhrNEygjlehHVICGt4RDAPC2EdEsIaRgTz8BHWISGsUQPBPE6EdUgIa+QQysgjrENiyl1W3RsighoYHMJ6aAhqYJAIawCIAGENABEgrIdGF5K4mAQMDqNBxoR69njRgIeEoXswIMDHgbAOCWENjwjxYSGsQ0JYo2UEeLwI65AQ1ugBAR4HwjokhDUCQYCHhaAODWGNgPUZ4IQVwkJYIzJdBThhjbAQ1hgI3yFOWCMshDUGrEmAE9YIC2GNkakb4IQ1wkJYA6UBTlgjLIQ1AETAlLusugcAESCsASAChDUARICwBoAIENYAEAHCGgAiQFgDQAQIawCIAGENABFgBiNa894ty8lti8vJWzZtTnZv2JD+Nkmeefls8rvzLyePra1OHyM8v73+9emj+h5ZO5F88cSR9CfMwHRzhOGGDRuTr+zYNQ3peR4/fTL53LFD6U8IBWHdKsLaRRc7o3qWX9q+K/2pvph3egX1gzv3JG/YuCn9zXw/O3Mq+cTRA+lPCAFh3SpT7lKzhlf3LG9zCmp51+LStGEDUI2whleqUVtY/z9gLAhreJW/kOjC+v8BY0FYA0AECGt4deDChfSRGw3lA1CNsIZXj5w6kT5y89TZ0+kjAGUIa3j17ZPHnSe6aKz1j0+vpT8BKENYw7v7jh6cBnAdGpvLpBhgPsIa3r1w4fw0gD9z7OB0wsvvC/Vo9bwV5h889CKTKICaCGu0RqUNzUy8/eDzyU0v/uHKl0JaYc66IEB9hDUARIC1QRwMbW0QrePxnsk2vn1hMblx48KrJqaodPHc5EujNL53ajX9bdjesXlL8tbJa9F0d72WGyffl9dd3R/Ra1u9dHE6zHD/hfPTHn7TC5xatOqdi0tXnnvv+o1XTfRZmzyn3lM9r7bh1+fOJL/oYBTMB5ZWklsm7422J7+4lrbjufPnkqcn2/GTyetX+SrTx9og2ra3bV5Mbk5XaSwuBJa9fxrmqfevuM0RYSEnF7GGtfVvqvyQ99ltO5M7l7aWhlmRDuqvrh6pFWi+tq8uHdDvX15Jbl24HEZWapT2Hdmf/lSPGrv3TYLwji1bzc+t9/bJM2vT99Z3WcjlM1YQPn7q5JX9rsuw1j5z1/K2q8K5Dl0T+f6kM9FFo+cRCzlhPh0QP9r1munBUecgFgWRAvjDW69Jf9M/vY5v7NidPHrd9cmdDcIys7y+/qGgkFYQ/mDXDcm+yXvS5Ln1/+qz0Ov4wvbrpn+7KZ1h/Hz365w+Y/07/XvtGz62oQ5tp1639i1LUIsWAXtg5x5v713ICOsR0cHx0LVuy5fmfXJlx7QX1Dc1GnodOlC7ptf/8LXXOwVhXWp01AA0eY/13ii8rA2I9o1/vOba9Kf2fHyyL2k7rSFdpPdOS/NqHx8qwnok1Ov49KQ32DRgPrWyM33UD/Wm1Wj4Dso6FITqBTbtxc+i12U9i1FtWu9NU6pvt0mfoc5IfFNDo5teDDWwCeuR+NjKdnOPOk9B1Vc5RAd5H71p0Wv2EYR16blc3mcF1N9t6277rNr+DNXYqVMyxJIIYT0SOk30xeffqkunzX0FtcoSXQZ1Rs+p3nIdPs6a2qbGp4vPUJ0SdU6GhrCGMx0MXfZc1Gts47S5Dr3OPks/6i3Pe68Vgj7Omtqk2vRHOvwM1aHwVQ8PBWENky7rgvf22EvSbcrarFHPk53Wz3LX0rb0Ubg+NAlq156/liR494E/TYd13ntkv/PQRj3nkBDWMHljRz05lSD66iGpR6txyn1T6aDqPVCZpM/GpA5tu2v5I7vrfTbpReOotUxBcZ2ZWW7ZvDio2jVhPSLqmaiHop6KeixfXz06nQxhsaejg6BpfVyTJj5//PD0QNfrzr70+rXQ1EMnj0//jSanFGl2p7UOrL+Zvdf60mNNGrHSxJ8yTUZuaCJQcX8oex+ach2KqH3yW6vH0p9e7WGH91Cf3ZBGhhDWI6GDUEuXZjO91GPR2tP/Nvmy6KI3p16RNYzUA1MQaSEpTZcvnkLr9WvW4DcnAaV/o68iTcO3UAOgv5efVafHmt2nQLTQDM0y6j1aqDHRjM3i/nD34RfNDXiVWx0/wycn21Y1jdx16YObF4ZTtyasA3Tgov/eje7gUnYA6AD1fXD6Yu0VqWH6aC6IrCwNhZ5bDUAVvd8KSldqHIulEL0/lp6/trGscRLtI9YGvIy22fXi57PnZtemXXr/XZXrukBYB+ilil5FE1r0pooWxwmRtVdU1TC5sDYUP6xx0wXrLczevGkhfXTZnxV+rmverdd8NuDFba7j5Jznfuli/c9WC2oNBWE9ErPCa15Ppi+WXpFCRmHT1DaHtULy/u/8ufRRNWuPv/h+7F5vK0X96uyZ9FE1Xw245TPUDE4tJFX15XLBOfSLry4Iawcui/2gH75C5k0bbb3W39Ro+Jr2+jPW/dH36n7oxmjTx3Ka59pL2DugYUN9cOlBZXydJSyts60e7CuI67D0WrsO6iFd4OvbaMPa0gNTeLiM29Qi6hZ1emcAxmW8PeuLtgsoGntbx+VhZ7ZhVV32zobG0tssYx2RE/q43hXjuHH0b7SfnG4NZKGpvXV615qmbBlWRT3xFS6z1TK6PZkP1hE5dS5MmockFhoQlyFsmbrD6HyNorBsI8qNNqytoairy1qcfVZga2EdLU5vYW1Ehkj3SXSlz6fJ4v2ZE8Yzrzpjs3WfSItiA6L7R1rMayxU7vM1isK6jbjaaMO6yQ1SdUDqbiEK5XxoKySyxfGtQh1G1wfre6FV8lyuLZTR8DrLReiqmYZ5umejRXGfPXXJdhvUv57TmPlcAMlSTtLyANk0fR9fQzHasBbr5ARRz0Oh/MTu114Z/6nxoU3X643lTuJdsJ5l6LPRLZ4so0nynjE0FnpuNdhVtC63pddadiZYZ0x3mdsm+2hV71q/97nmdJ0x3UV3T85Kmza2QzTqsNYt+ENimYY8ZGq4rDPpVJvVzVgVnFqZrnjwK8h1JqTw1L8pC1jr/qGwe2jSWOTLMQpBPYd1Xe5floSe9exQ11J0+6v8maHeD90EWL/3SY2Ma91an52lsdVryT7PIbINJr3Mdg4WGN0F2ld9rimtAjfvAFQAqAfvatbpoA5S1xq7DkKtZFfke/t04Pns6VUpez06+HUDW+vKe76owfqbgy+UjhLS3chDufGAVhXUYlVFuvO4dfVEdWCenXw2//vyuatmfqoB1AVdhbrGc2fhXrVvBsSUu6PuWcu8dRK6oh2sSR19qL7jcVEhVwrHx0/NX+ujbdqGquGcMZyNPbZmL+2poVa5UXdCz08515d+p46BOhquvfAYjT6stY6EZYiYbw9UrN87dmrEmqwD3dR3J89tLcX4oOfWNlT5jwaloq7oM9TNBNDM6MNavjw5detzh1cYNV3Oc8h0at1Xg6oe7ddO2Nag9kHPPWuSVCi9/3l0MwHGXDdDWE8oKH2u4etCp7FldT682t8fO9Rbg6oLnbqhQNd0o4I6o4PU8w49CNWo3H+8v89wCAjrlMoh1rt4WCmoqxaBx6vpVFp3uunrYNcNBbosx2hfrLvUaxaEvrRVslCnqM/PMHaEdY4ODt0Kqoteig58gtqNDnaNilBw9yG7LVebYaO/rUkhrmty673x0dlQB0I3qm1LFtiURNwR1gXamXQfurZ6UQoaDSui9GGjXqTeP5Ul+uihKUT3Hd7fyigM/U01RtaJUU07G+pRd9GBaPsYGyrCuoQCQWGqUNAO1bQXoFDRgagDSX+zr57hkKgsoWBTaPv4fMomnVTR56dQ02epz7VJo6H/VyGpv6W/qX2viXwQ1t0uvR6N8W+zR13k+xgTvQ7tD+q5D9HoJ8XUpQH4WoBHExA0iUarkpVNptGoBS1ApJ1Pj3997sz0AEK7ss/npk0L0zuoVI27zRpKTWU/OPmMfH0+mgyk59QSrXr+G/W9MJlG4al11LU8r57/vyfP2+a+oUk9WtJX66pre/KTZ7RvaluePLMWzPj+4jFW9h5K/hjTQlH6TLUGfNOGrkOm3CWsAaBbzGAEgKEirAEgAoQ1AESgSc1aqFsDQH3mzKVnDQARIKwBIAJNyyBCKQQA5muUt/SsASACPnrWQu8aAKo1zlpfYS0ENgBczUvOUgYBgPZ46xD7DGufvXQAQE5bAUtJBMCYec/WNnvDBDaAMWolV9suXRDYAMai1TxtO6wzhDaAIWs9S7sK6zyCG0Ds+shOAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQryT5fw2Jdb/hdIH+AAAAAElFTkSuQmCC"

class Krutidev2unicode(QMainWindow):
    def __init__(self):
        super().__init__()

        self.columns_to_process = []
        self.file_path = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Unicode-Krutidev Excel Converter")
        self.setGeometry(400, 450, 500, 200)

        centerOnScreen(self)

        # Logo
        pixmap = QPixmap()
        pixmap.loadFromData(QByteArray.fromBase64(base64_image))
        self.setWindowIcon(QIcon(pixmap))

        layout = QVBoxLayout()

        #-----textbox
        self.text_box = QTextEdit(self)
        self.text_box.setFixedHeight(40) 
        layout.addWidget(self.text_box)

        load_button = QPushButton("Load Excel", self)
        load_button.clicked.connect(self.load_excel)
        layout.addWidget(load_button)

        
        self.table_widget = QTableWidget(self)
        layout.addWidget(self.table_widget)

        horizontal_layout = QHBoxLayout()

        krutidev_button = QPushButton("Unicode to KrutiDev", self)
        krutidev_button.clicked.connect(self.convert_unicode)
        horizontal_layout.addWidget(krutidev_button)

        unicode_button = QPushButton("KrutiDev to Unicode", self)
        unicode_button.clicked.connect(self.convert_krutidev)
        horizontal_layout.addWidget(unicode_button)

        layout.addLayout(horizontal_layout)

        # processing message
        self.processing_label = QLabel()
        self.processing_label.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.processing_label.setText("<p style='color: blue; backgroundcolor: white; font-size: 16px'>Please wait... The conversion is in process...</p>")
        self.processing_label.setHidden(True)
        layout.addWidget(self.processing_label)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.central_widget.setLayout(layout)

        self.setStyleSheet("""
            QMainWindow {
                background-color:#f5f5f5; /* White background */
            }
            QLineEdit {
                background-color: white;
                border: 1px solid #ccc; /* Light Gray border */
                padding: 4px;
                width: 150px; /* Adjust as needed */
            }
            QPushButton {
                background-color: #0d6efd; /* Dark Blue button */
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 3px;
                width: 150px;
            }
            QPushButton:hover {
                background-color: #0951ba; /* Slightly darker grey on hover */
            }
        """)

    def load_excel(self):
        file_dialog = QFileDialog()
        self.file_path, _ = file_dialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)")
        self.text_box.setPlainText(self.file_path)

        if self.file_path:
            self.load_headers()

    def load_headers(self):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        header_row = sheet[1]
        headers = [cell.value for cell in header_row]

        self.table_widget.setColumnCount(len(headers))
        self.table_widget.setRowCount(1)
        # self.table_widget.setHorizontalHeaderLabels(headers)

        for col in range(len(headers)):
            checkbox = QCheckBox(headers[col], self)
            self.table_widget.setCellWidget(0, col, checkbox)

        # Adjust column widths based on content
        for col in range(len(headers)):
            self.table_widget.resizeColumnToContents(col)

    def clear_headers(self):
        try:
            # Remove all widgets from the first row of the table (headers row)
            for col in range(self.table_widget.columnCount()):
                widget = self.table_widget.cellWidget(0, col)
                if widget is not None:
                    widget.setParent(None)

            # Clear the horizontal header labels
            self.table_widget.setHorizontalHeaderLabels([])

            # Clear the row count (excluding header row)
            self.table_widget.setRowCount(0)

            # Clear the column count
            self.table_widget.setColumnCount(0)

        except Exception as e:
            print(f"An error occurred while clearing headers: {e}")


    def convert_krutidev(self):
        self.process_excel(KrutiDev_to_Unicode)

    def convert_unicode(self):
        self.process_excel(Unicode_to_KrutiDev)

    def process_excel(self, conversion_function):
        if not self.file_path:
            QMessageBox.warning(self, 'File missing', 'Please select an EXCEL file.')
            return
        
        self.columns_to_process = []
        for col in range(self.table_widget.columnCount()):
            checkbox = self.table_widget.cellWidget(0, col)
            if checkbox.isChecked():
                self.columns_to_process.append(col + 1)

        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active

        self.processing_label.setVisible(True)
        show_progress("Please wait... The conversion is in process.")

        for row in range(2, sheet.max_row + 1):
            for col in self.columns_to_process:
                cell = sheet.cell(row=row, column=col)
                krutidev_text = cell.value

                # Skip processing if the cell value is None
                if krutidev_text is not None:
                    # print(f'---cell--- Row: {row}, Column: {col}, Value: {krutidev_text}')

                    # Apply the selected conversion function
                    modified_text = conversion_function(krutidev_text)

                    # Update the cell value with the modified text
                    cell.value = modified_text


        self.processing_label.setHidden(True)

        file_dialog = QFileDialog()
        output_file_path, _ = file_dialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if output_file_path:
            wb.save(output_file_path)

        message = f'Conversion Successfull'
        QMessageBox.information(None, "Success", message)

        subprocess.Popen(['start', '', output_file_path], shell=True)

        # Clear headers
        self.clear_headers()

def centerOnScreen(self):
    # Get the screen geometry
    screen_geometry = QGuiApplication.primaryScreen().geometry()

    # Calculate the center position
    x = (screen_geometry.width() - self.width()) // 2
    y = (screen_geometry.height() - self.height()) // 2

    # Move the window to the center position
    self.move(x, y)

def KrutiDev_to_Unicode(krutidev_substring):

    if krutidev_substring is None:
        return None
    
    modified_substring = krutidev_substring
    
    array_one = ["ñ","Q+Z","sas","aa",")Z","ZZ","‘","’","“","”",
    
    "å",  "ƒ",  "„",   "…",   "†",   "‡",   "ˆ",   "‰",   "Š",   "‹", 
    
    "¶+",   "d+", "[+k","[+", "x+",  "T+",  "t+", "M+", "<+", "Q+", ";+", "j+", "u+",
    "Ùk", "Ù", "Dr", "–", "—","é","™","=kk","f=k",  
    
    "à",   "á",    "â",   "ã",   "ºz",  "º",   "í", "{k", "{", "=",  "«",   
    "Nî",   "Vî",    "Bî",   "Mî",   "<î", "|", "K", "}",
    "J",   "Vª",   "Mª",  "<ªª",  "Nª",   "Ø",  "Ý", "nzZ",  "æ", "ç", "Á", "xz", "#", ":",
    
    "v‚","vks",  "vkS",  "vk",    "v",  "b±", "Ã",  "bZ",  "b",  "m",  "Å",  ",s",  ",",   "_",
    
    "ô",  "d", "Dk", "D", "[k", "[", "x","Xk", "X", "Ä", "?k", "?",   "³", 
    "pkS",  "p", "Pk", "P",  "N",  "t", "Tk", "T",  ">", "÷", "¥",
    
    "ê",  "ë",   "V",  "B",   "ì",   "ï", "M+", "<+", "M",  "<", ".k", ".",    
    "r",  "Rk", "R",   "Fk", "F",  ")", "n", "/k", "èk",  "/", "Ë", "è", "u", "Uk", "U",   
    
    "i",  "Ik", "I",   "Q",    "¶",  "c", "Ck",  "C",  "Hk",  "H", "e", "Ek",  "E",
    ";",  "¸",   "j",    "y", "Yk",  "Y",  "G",  "o", "Ok", "O",
    "'k", "'",   "\"k",  "\"",  "l", "Lk",  "L",   "g", 
    
    "È", "z", 
    "Ì", "Í", "Î",  "Ï",  "Ñ",  "Ò",  "Ó",  "Ô",   "Ö",  "Ø",  "Ù","Ük", "Ü",
    
    "‚",    "ks",   "kS",   "k",  "h",    "q",   "w",   "`",    "s",    "S",
    "a",    "¡",    "%",     "W",  "•", "·", "∙", "·", "~j",  "~", "\\","+"," ः",
    "^", "*",  "Þ", "ß", "(", "¼", "½", "¿", "À", "¾", "A", "-", "&", "&", "Œ", "]","~ ","@"]
    
    array_two = ["॰","QZ+","sa","a","र्द्ध","Z","\"","\"","'","'",
    
    "०",  "१",  "२",  "३",     "४",   "५",  "६",   "७",   "८",   "९",   
    
    "फ़्",  "क़",  "ख़", "ख़्",  "ग़", "ज़्", "ज़",  "ड़",  "ढ़",   "फ़",  "य़",  "ऱ",  "ऩ",    
    "त्त", "त्त्", "क्त",  "दृ",  "कृ","न्न","न्न्","=k","f=",
    
    "ह्न",  "ह्य",  "हृ",  "ह्म",  "ह्र",  "ह्",   "द्द",  "क्ष", "क्ष्", "त्र", "त्र्", 
    "छ्य",  "ट्य",  "ठ्य",  "ड्य",  "ढ्य", "द्य", "ज्ञ", "द्व",
    "श्र",  "ट्र",    "ड्र",    "ढ्र",    "छ्र",   "क्र",  "फ्र", "र्द्र",  "द्र",   "प्र", "प्र",  "ग्र", "रु",  "रू",
    
    "ऑ",   "ओ",  "औ",  "आ",   "अ", "ईं", "ई",  "ई",   "इ",  "उ",   "ऊ",  "ऐ",  "ए", "ऋ",
    
    "क्क", "क", "क", "क्", "ख", "ख्", "ग", "ग", "ग्", "घ", "घ", "घ्", "ङ",
    "चै",  "च", "च", "च्", "छ", "ज", "ज", "ज्",  "झ",  "झ्", "ञ",
    
    "ट्ट",   "ट्ठ",   "ट",   "ठ",   "ड्ड",   "ड्ढ",  "ड़", "ढ़", "ड",   "ढ", "ण", "ण्",   
    "त", "त", "त्", "थ", "थ्",  "द्ध",  "द", "ध", "ध", "ध्", "ध्", "ध्", "न", "न", "न्",    
    
    "प", "प", "प्",  "फ", "फ्",  "ब", "ब", "ब्",  "भ", "भ्",  "म",  "म", "म्",  
    "य", "य्",  "र", "ल", "ल", "ल्",  "ळ",  "व", "व", "व्",   
    "श", "श्",  "ष", "ष्", "स", "स", "स्", "ह", 
    
    "ीं", "्र",    
    "द्द", "ट्ट","ट्ठ","ड्ड","कृ","भ","्य","ड्ढ","झ्","क्र","त्त्","श","श्",
    
    "ॉ",  "ो",   "ौ",   "ा",   "ी",   "ु",   "ू",   "ृ",   "े",   "ै",
    "ं",   "ँ",   "ः",   "ॅ",  "ऽ", "ऽ", "ऽ", "ऽ", "्र",  "्", "?", "़",":",
    "‘",   "’",   "“",   "”",  ";",  "(",    ")",   "{",    "}",   "=", "।", ".", "-",  "µ", "॰", ",","् ","/"]
    
    array_one_length = len(array_one)
    
    # Specialty characters
    
    # Move "f"  to correct position and replace
    modified_substring = "  " + modified_substring + "  "
    position_of_f = modified_substring.rfind("f")              # vfHk"ksd cUnsokj 
    while (position_of_f != -1):         
        modified_substring = modified_substring[:position_of_f] + modified_substring[position_of_f+1] + modified_substring[position_of_f] +  modified_substring[position_of_f+2:]
        position_of_f = modified_substring.rfind("f",0, position_of_f - 1 ) # search for f ahead of the current position.
    modified_substring = modified_substring.replace("f","ि")
    modified_substring = modified_substring.strip()
    
    # Move "half R"  to correct position and replace
    modified_substring = "  " + modified_substring + "  "
    position_of_r = modified_substring.find("Z")
    set_of_matras =  ["‚",    "ks",   "kS",   "k",     "h",    "q",   "w",   "`",    "s",    "S", "a",    "¡",    "%",     "W",   "·",   "~ ", "~"]
    while (position_of_r != -1):    
        modified_substring = modified_substring.replace("Z","",1)
        if modified_substring[position_of_r - 1] in set_of_matras:
            modified_substring = modified_substring[:position_of_r - 2] + "j~" + modified_substring[position_of_r - 2:]
        else:
            modified_substring = modified_substring[:position_of_r - 1] + "j~" + modified_substring[position_of_r - 1:]
        position_of_r = modified_substring.find("Z")
    modified_substring = modified_substring.strip()
    
    # Replace ASCII with Unicode
    for input_symbol_idx in range(0, array_one_length):
        modified_substring = modified_substring.replace(array_one[input_symbol_idx ] , array_two[input_symbol_idx] )
    
    
    return modified_substring


def Unicode_to_KrutiDev(unicode_substring):

    if unicode_substring is None:
        return None
    
    modified_substring = unicode_substring
    
    array_one = ["‘",   "’",   "“",   "”",   "(",    ")",   "{",    "}",   "=", "।",  "?",  "-",  "µ", "॰", ",", ".", "् ", 
    "०",  "१",  "२",  "३",     "४",   "५",  "६",   "७",   "८",   "९", "x", 
    
    "फ़्",  "क़",  "ख़",  "ग़", "ज़्", "ज़",  "ड़",  "ढ़",   "फ़",  "य़",  "ऱ",  "ऩ",  
    "त्त्",   "त्त",     "क्त",  "दृ",  "कृ",
    
    "ह्न",  "ह्य",  "हृ",  "ह्म",  "ह्र",  "ह्",   "द्द",  "क्ष्", "क्ष", "त्र्", "त्र","ज्ञ",
    "छ्य",  "ट्य",  "ठ्य",  "ड्य",  "ढ्य", "द्य","द्व",
    "श्र",  "ट्र",    "ड्र",    "ढ्र",    "छ्र",   "क्र",  "फ्र",  "द्र",   "प्र",   "ग्र", "रु",  "रू",
    "्र",
    
    "ओ",  "औ",  "आ",   "अ",   "ई",   "इ",  "उ",   "ऊ",  "ऐ",  "ए", "ऋ",
    
    "क्",  "क",  "क्क",  "ख्",   "ख",    "ग्",   "ग",  "घ्",  "घ",    "ङ",
    "चै",   "च्",   "च",   "छ",  "ज्", "ज",   "झ्",  "झ",   "ञ",
    
    "ट्ट",   "ट्ठ",   "ट",   "ठ",   "ड्ड",   "ड्ढ",  "ड",   "ढ",  "ण्", "ण",  
    "त्",  "त",  "थ्", "थ",  "द्ध",  "द", "ध्", "ध",  "न्",  "न",  
    
    "प्",  "प",  "फ्", "फ",  "ब्",  "ब", "भ्",  "भ",  "म्",  "म",
    "य्",  "य",  "र",  "ल्", "ल",  "ळ",  "व्",  "व", 
    "श्", "श",  "ष्", "ष",  "स्",   "स",   "ह",     
    
    "ऑ",   "ॉ",  "ो",   "ौ",   "ा",   "ी",   "ु",   "ू",   "ृ",   "े",   "ै",
    "ं",   "ँ",   "ः",   "ॅ",    "ऽ",  "् ", "्" ]
    
    array_two = ["^", "*",  "Þ", "ß", "¼", "½", "¿", "À", "¾", "A", "\\", "&", "&", "Œ", "]","-","~ ", 
    "å",  "ƒ",  "„",   "…",   "†",   "‡",   "ˆ",   "‰",   "Š",   "‹","Û",
    
    "¶",   "d",    "[k",  "x",  "T",  "t",   "M+", "<+", "Q",  ";",    "j",   "u",
    "Ù",   "Ùk",   "Dr",    "–",   "—",       
    
    "à",   "á",    "â",   "ã",   "ºz",  "º",   "í", "{", "{k",  "«", "=","K", 
    "Nî",   "Vî",    "Bî",   "Mî",   "<î", "|","}",
    "J",   "Vª",   "Mª",  "<ªª",  "Nª",   "Ø",  "Ý",   "æ", "ç", "xz", "#", ":",
    "z",
    
    "vks",  "vkS",  "vk",    "v",   "bZ",  "b",  "m",  "Å",  ",s",  ",",   "_",
    
    "D",  "d",    "ô",     "[",     "[k",    "X",   "x",  "?",    "?k",   "³", 
    "pkS",  "P",    "p",  "N",   "T",    "t",   "÷",  ">",   "¥",
    
    "ê",      "ë",      "V",  "B",   "ì",       "ï",     "M",  "<",  ".", ".k",   
    "R",  "r",   "F", "Fk",  ")",    "n", "/",  "/k",  "U", "u",   
    
    "I",  "i",   "¶", "Q",   "C",  "c",  "H",  "Hk", "E",   "e",
    "¸",   ";",    "j",  "Y",   "y",  "G",  "O",  "o",
    "'", "'k",  "\"", "\"k", "L",   "l",   "g",      
    
    "v‚",    "‚",    "ks",   "kS",   "k",     "h",    "q",   "w",   "`",    "s",    "S",
    "a",    "¡",    "%",     "W",   "·",   "~ ", "~"]
    
    array_one_length = len(array_one)
    
    # Specialty characters
    modified_substring = modified_substring.replace ("क़", "क़")   
    modified_substring = modified_substring.replace ("ख़‌", "ख़")
    modified_substring = modified_substring.replace ("ग़", "ग़")
    modified_substring = modified_substring.replace ("ज़", "ज़")
    modified_substring = modified_substring.replace ("ड़", "ड़")
    modified_substring = modified_substring.replace ("ढ़", "ढ़")
    modified_substring = modified_substring.replace ("ऩ", "ऩ")
    modified_substring = modified_substring.replace ("फ़", "फ़")
    modified_substring = modified_substring.replace ("य़", "य़")
    modified_substring = modified_substring.replace ("ऱ", "ऱ")
    modified_substring = modified_substring.replace("ि","f")
    
    # Replace Unicode with ASCII
    for input_symbol_idx in range(0, array_one_length):
        modified_substring = modified_substring.replace(array_one[input_symbol_idx ] , array_two[input_symbol_idx] )
    
    # Move "f"  to correct position
    modified_substring = "  " + modified_substring + "  "
    position_of_f = modified_substring.find("f")
    while (position_of_f != -1):    
        modified_substring = modified_substring[:position_of_f-1] + modified_substring[position_of_f] + modified_substring[position_of_f-1] + modified_substring[position_of_f+1:]
        position_of_f = modified_substring.find("f", position_of_f +1 ) # search for f ahead of the current position.
    modified_substring = modified_substring.strip()
    
    # Move "half R"  to correct position and replace
    modified_substring = "  " + modified_substring + "  "
    position_of_r = modified_substring.find("j~")
    set_of_matras =  ["‚",    "ks",   "kS",   "k",     "h",    "q",   "w",   "`",    "s",    "S", "a",    "¡",    "%",     "W",   "·",   "~ ", "~"]
    while (position_of_r != -1):    
        modified_substring = modified_substring.replace("j~","",1)
        if modified_substring[position_of_r + 1] in set_of_matras:
            modified_substring = modified_substring[:position_of_r + 2] + "Z" + modified_substring[position_of_r + 2:]
        else:
            modified_substring = modified_substring[:position_of_r + 1] + "Z" + modified_substring[position_of_r + 1:]
        position_of_r = modified_substring.find("j~")
    modified_substring = modified_substring.strip()
    
    return modified_substring

def show_progress(message):
    success_msg_box = QMessageBox()
    success_msg_box.setIcon(QMessageBox.Information)
    success_msg_box.setWindowTitle("Processing")
    success_msg_box.setText(message)
    success_msg_box.setStandardButtons(QMessageBox.NoButton)
    pixmap = QPixmap()
    pixmap.loadFromData(QByteArray.fromBase64(base64_image)) 
    success_msg_box.setWindowIcon(QIcon(pixmap))
    
    # Timer to close the message box after a specified time (in milliseconds)
    close_timer = QTimer()
    close_timer.timeout.connect(success_msg_box.accept)
    close_timer.start(2000)  

    success_msg_box.exec()



if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = Krutidev2unicode()
    window.show()
    sys.exit(app.exec_())

